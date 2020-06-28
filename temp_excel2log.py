# -*- coding: utf-8 -*-
"""
Created on Wed Jun 24 19:43:16 2020

@author: Brian Glassman
"""

xlfile = 'ref_Budget Buckets.xlsm'
logfile = 'raw_log.txt'

if __name__ == '__main__':

    import openpyxl


    """Script for converting from my original Excel file to the new log format
    Purely for testing and rapid development, shouldn't be used in final version
    Note: creates the "raw" log"""

    # NOTE: HAVE TO BE CAREFUL ABOUT INDEXING. sheet.cell() uses 1-index, but
    # tuples created from rows/columns use 0-index



    # Keys that will be in the production-version log file
    transaction_keys = ('Status', 'Idk', 'Date', 'Custom Description',
                        'Auto Description', 'Auto Category', 'Raw Amount')
    n = len(transaction_keys)
    # Keys only used for these temporary versions
    temp_keys = ('_', '_', 'Split_raw', '_', 'My Category')

    #%% Read in transactions from file
    log = []
    wb = openpyxl.load_workbook(filename=xlfile, read_only=False, data_only=True)
    try:
        # Different years are stored in different sheets
        sheets = ['Log 2020']
        for ws in (wb[sheetname] for sheetname in sheets):
            # Each row after the header is one transaction
            # Skip header in row 1
            # Columns A-G (1-7) are original data, columns H+ are post-processed
            # Columns J (10) and L (12) are used to categorize and split
            #   Only needed for these temporary versions
            for raw_row in ws.iter_rows(min_row=2, max_row=20, max_col=12, values_only=True):
                transaction = {k:v for k,v in zip(transaction_keys, raw_row)}
                transaction['Date'] = transaction['Date'].date()
                transaction['temp'] = {k:v for k,v in zip(temp_keys, raw_row[n:]) if k != '_'}
                log.append(transaction)
    finally:
        # Openpyxl doesn't release open read-only workbooks
        wb.close()

    #%% Create the JSON log
    import json

    # Should be in chronological order (earliest at the top, latest at the bottom)
    log = sorted(log, key=lambda x: x['Date'])

    for num, transaction in enumerate(log):
        del transaction['Status'] # TODO will be needed once predictive stuff is in use
        del transaction['Idk'] # Venmo fills this column, but that will be taken care of separately
        transaction['Raw ID'] = num # Unique ID number for each transaction

    jstr = json.dumps(log, default=str, indent=2)

    print(jstr)

    json.dump(log, open(logfile, 'w'), indent=2, default=str)

    #%% NOTES ON LOG FORMAT
    """
    Save as a chronological list, rather than dictionary keyed by date
        Once read into memory, can pivot to dict keyed by date, category, etc.
    """
