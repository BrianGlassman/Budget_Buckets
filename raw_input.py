# -*- coding: utf-8 -*-
"""
Created on Sat Aug  1 16:32:23 2020

@author: Brian Glassman


#%%
Parses raw input files from banks, Venmo, etc.

Parsers are classes to make it easier to group information and helper methods
"""

#%% RawTransaction definition
# TODO move to library
import datetime
import dateutil
from dataclasses import dataclass
import re # Needed for isinstance check

from templates import templates

@dataclass
class RawTransaction:
    """Transactions as read in from raw input files."""

    date: datetime.date
    value: float
    desc: str
    auto_cat: str
    account: str

    def __init__(self, date, value, desc=None, auto_cat=None, account=None):
        self.date = self.__make_date(date)
        self.value = float(value)
        self.desc = desc if desc is not None else ""
        self.auto_cat = auto_cat
        self.account = account

    def __make_date(self, raw):
        """Convert input to the appropriate type"""

        # Ensure date is a consistent type to avoid many, many headaches
        # Can't use isinstance, because datetime.datetime is a subclass of datetime.date
        if type(raw) is datetime.date:
            return raw
        elif isinstance(raw, datetime.date):
            return raw.date()
        else:
            return dateutil.parser.parse(raw).date()

    def to_dict(self):
        """Return a dictionary of the contained values"""
        return self.__dict__

    # Helper functions to check different fields
    def __check_desc(self, pattern):
        """Assumes that pattern contains a desc field"""
        mask = pattern['desc']
        desc = self.desc
        if isinstance(mask, re.Pattern):
            # Regex matching (None if no match found)
            return mask.search(desc) is not None
        elif isinstance(mask, str):
            # Exact string matching
            return desc == mask
        else:
            raise NotImplementedError("Unknown mask type")
    def __check_value(self, pattern):
        """Assumes that pattern contains a value field"""
        mask = pattern['value']
        if isinstance(mask, (int, float)):
            # Exact value matching
            return self.value == mask
        else:
            # Range matching
            assert len(mask) == 2
            assert mask[0] < mask[1]
            return mask[0] <= self.value <= mask[1]

    # Mapping from field name to helper function
    _checker = {'desc': __check_desc,
                'value': __check_value,
                }

    def match_templates(self):
        """Check self against the common templates. Return whichever template
        matches, or None if no match"""

        matched = None
        # Check the transaction against all templates in order
        for template in templates:
            old = template['old']
            match = True
            for key in old:
                # Run the checker for each field that has a pattern
                match = match and self._checker[key](self, old)
                if not match: break # Exit loop if any pattern fails

            if match:
                # Template matched, stop searching
                matched = template
                break
        return matched

    def apply_template(self, template):
        """Convert self from RawTransaction to one or more Transactions as
        specified by the template"""

        new = template['new']
        if isinstance(new, dict):
            # One-to-one conversion, make a list for iterating
            new = [new]
        else:
            # Assume one-to-many conversion, already a list
            pass

        ret = []
        for n in new:
            # TODO use alternate constructor once implemented
            #   Or some other smart way of converting RawTransaction to Transaction
            # Deep copy n to avoid changing the base common templates
            n = dict(n)
            n['recurrence'] = n.pop('split')

            d = self.to_dict()
            d.update(n)
            # Keep only keys that are used for Transaction
            d = {k:v for k,v in d.items() if k in Transaction.fields()}
            ret.append(Transaction(**d))
        return ret

def test_checker():
    #TODO test splitting transaction
    #TODO test trivial cases
    #TODO test value range matching
    return True

def test_RawTransaction():
    # TODO
    # date as string
    # date as datetime.datetime
    # date as datetime.date
    # empty description
	# value as int
	# value as float
	# value as negative int
	# value as negative float
    # value as string like "##.##" <-- normal case float
    # value as string like "##" <-- normal case int
    # value as string like "-##.####" <-- normal case negative float
    # value as string like "-##" <-- normal case negative int
    # value as string like "$##.##"
    # should NOT test any weirder values. That's the parser's job to take care of
    return True

#%% Transaction definition (TO BE MOVED ELSEWHERE)
# TODO move to library
import datetime
import dataclasses
from dataclasses import dataclass
@dataclass
class Transaction:
    """Transactions as used internally (has been split, categorized, etc.)"""

    date: datetime.date
    value: float
    desc: str
    category: str
    recurrence: str
    # TODO add raw_id and mod_id

    # TODO alternate constructor for creating Transaction out of RawTransaction

    def __init__(self, date, value, desc=None, category=None, recurrence=None):
        self.date = self.__make_date(date)
        self.value = float(value)
        self.desc = desc if desc is not None else ""
        self.category = category
        self.recurrence = recurrence

    def __make_date(self, raw):
        """Convert input to the appropriate type"""

        # Ensure date is a consistent type to avoid many, many headaches
        # Can't use isinstance, because datetime.datetime is a subclass of datetime.date
        if type(raw) is datetime.date:
            return raw
        elif isinstance(raw, datetime.date):
            return raw.date()
        else:
            return dateutil.parser.parse(raw).date()

    @staticmethod
    def fields():
        """Return a list of field names as strings"""
        return [x.name for x in dataclasses.fields(Transaction)]

#%% Definitions
import csv
from abc import ABC


def parse(infile, parser, *args, **kwargs):
    """Run the provided parsing logic on the given file."""
    return parser.parse(infile, *args, **kwargs)


class USAA_parser(ABC):
    """Namespace for handling csv output files from USAA.

    Files have no header, csv with lines of the form:
    [string status],,[string MM/DD/YYYY],,[string description],[string category],[float value]

    Note: value is -[float] for expenses, --[float] for income
    """

    fieldnames = ['status', 'idk', 'date', 'custom_desc', 'auto_desc', 'auto_cat', 'value']

    @classmethod
    def parse(cls, infile, account):
        """Parse the given file, tagging transactions as the given account.

        Args:
            infile (file path): path to the input file to open
            account (string): the account to tag transactions as
        """
        assert account is not None, "USAA parser called without specifying account"
        with open(infile, 'r') as f:
            return [cls.make_transaction(line, account) for line in csv.DictReader(f, cls.fieldnames)]

    @classmethod
    def make_transaction(cls, line, account):
        """Convert a line from the input file into a RawTransaction object.

        Args:
            line (dict of strings): the line read in from the input file
            account (string): the account to tag transactions as
        Returns:
            raw_transaction object
        """
        # Make empty fields None
        line = {k:(v if v else None) for k,v in line.items()}
        desc = cls.get_desc(**line)
        value = cls.format_value(**line)
        return RawTransaction(date = line['date'],
                              desc = desc,
                              auto_cat = line['auto_cat'],
                              value = value,
                              account = account)

    @staticmethod
    def get_desc(custom_desc, auto_desc, **excess):
        """Combines custom_desc and auto_desc
        Args:
            custom_desc (str): manually input description
            auto_desc (str): automatically generated description
            **excess: unused fields from the input line
        Returns:
            one desc filled: return unchanged
            both filled: "[auto] --- [custom]"
            neither: empty string
        """
        if custom_desc and auto_desc:
            return f"{auto_desc} --- {custom_desc}"
        elif custom_desc and (not auto_desc): return custom_desc
        elif (not custom_desc) and auto_desc: return auto_desc
        elif (not custom_desc) and (not auto_desc): return ""
        else: raise RuntimeError # Shouldn't be possible

    @staticmethod
    def format_value(value, **excess):
        """Properly format the value
        Args:
            value (str): monetary value. Positive values may have "--" prefix
            **excess: unused fields from the input line
        Returns:
            value as a float
        """
        if value.startswith('--'):
            return float(value[2:])
        else:
            return float(value)


def venmo_parser():
    """
    Two line header of the form:
     Username,ID,Datetime,Type,Status,Note,From,To,Amount (total),Amount (fee),Funding Source,Destination,Beginning Balance,Ending Balance,Statement Period Venmo Fees,Year to Date Venmo Fees,Disclaimer
    [string],,,,,,,,,,,,$[float],,,,

    Data lines of the form:
    ,[int],[YYYY-MM-DDTHH:MM:SS],[string type],[string status],[string description],[string payer],[string payee],[+-] $[float],,[string funding source],,,,,,

    Four line footer of the form:
    ,,,,,,,,,,,,,$[float],,,
    ,,,,,,,,,,,,,,$[float],,
    ,,,,,,,,,,,,,,,$[float],
    ,,,,,,,,,,,,,,,,"[multi-line text disclaimer]"

    """


def cash_parser():
    pass


class excel_parser(ABC):
    """Namespace for parsing my existing Excel files"""

    xlfile = 'ref_Budget Buckets.xlsm'

    fieldnames = {'Date': 'date',
                  'Custom Description': 'custom_desc',
                  'Auto Description': 'auto_desc',
                  'USAA Category': 'auto_cat',
                  'Raw Amount': 'value',
                  'Account': 'account',}

    @classmethod
    def parse(cls, infile=xlfile, accounts=None):
        """Parse the file.

        Args:
            infile (file path): path to the input file to open
        """
        import openpyxl
        wb = openpyxl.load_workbook(filename=cls.xlfile, read_only=False, data_only=True)

        out = []
        try:
            # Different years are stored in different sheets
            sheets = ['Log 2018']
            for ws in (wb[sheetname] for sheetname in sheets):
                # Each row after the header is one transaction
                # Skip header in row 1

                # Format as {field : 0-indexed column number}, only keep used columns
                header = {cell.value:c for c,cell in enumerate(ws[1]) if cell.value in cls.fieldnames.keys()}
                # no min_col because that would mess up the indexing
                max_col = max(header.values()) + 1  # max_col parameter uses 1-index

                for raw_row in ws.iter_rows(min_row=2, max_col=max_col, values_only=True):
                    row = {cls.fieldnames[field]:raw_row[c] for field,c in header.items()}

                    # Filter by account
                    if accounts is not None and row['account'] in accounts:
                        out.append(cls.make_transaction(row))
        finally:
            # Openpyxl doesn't release open read-only workbooks
            wb.close()

        return out


    @classmethod
    def make_transaction(cls, line):
        """Convert a line from the input file into a RawTransaction object.

        Args:
            line (dict of strings): the line read in from the input file
        Returns:
            raw_transaction object
        """
        desc = cls.get_desc(**line)
        value = cls.format_value(**line)
        return RawTransaction(date = line['date'],
                              desc = desc,
                              auto_cat = line['auto_cat'],
                              value = value,
                              account = line['account'])

    @staticmethod
    def get_desc(custom_desc, auto_desc, **excess):
        """Combines custom_desc and auto_desc
        Args:
            custom_desc (str): manually input description
            auto_desc (str): automatically generated description
            **excess: unused fields from the input line
        Returns:
            one desc filled: return unchanged
            both filled: "[auto] --- [custom]"
            neither: empty string
        """
        if custom_desc and auto_desc:
            return f"{auto_desc} --- {custom_desc}"
        elif custom_desc and (not auto_desc): return custom_desc
        elif (not custom_desc) and auto_desc: return auto_desc
        elif (not custom_desc) and (not auto_desc): return ""
        else: raise RuntimeError  # Shouldn't be possible

    @staticmethod
    def format_value(value, **excess):
        """Properly format the value
        Args:
            value (str): monetary value. Positive values may have "--" prefix
            **excess: unused fields from the input line
        Returns:
            value as a float
        """
        if isinstance(value, str) and value.startswith('--'):
            return float(value[2:])
        else:
            return float(value)

class excel_cat_parser(excel_parser):
    fieldnames = {'Date': 'date',
                  'Custom Description': 'custom_desc',
                  'Auto Description': 'auto_desc',
                  'Raw Amount': 'value',
                  'Account': 'account',
                  'My Category': 'category',
                  'Split_raw': 'split'}

    @staticmethod
    def get_recurrence(split, **excess):
        # Apply recurrence
        if split == 1:
            # Single transaction
            return None
        else:
            # Recurring transaction
            # Have to do some fuzzy logic, this won't be in production
            # NOTE: initial count is reduced by one to account for first instance
            if split == 14:
                # Salary
                return '14d' # Apply over 14 days, don't normalize
            elif 28 <= split <= 31:
                # Monthly
                return '1m' # Apply over 1 month, don't normalize
            elif 365 <= split <= 366:
                # Yearly
                return '1y' # Apply over 1 year, don't normalize

    @classmethod
    def make_transaction(cls, line):
        """Convert a line from the input file into a Transaction object.

        Args:
            line (dict of strings): the line read in from the input file
        Returns:
            raw_transaction object
        """
        desc = cls.get_desc(**line)
        value = cls.format_value(**line)
        rec = cls.get_recurrence(**line)
        return Transaction(date = line['date'],
                           desc = desc,
                           category = line['category'],
                           value = value,
                           recurrence = rec)

#%% Testing
def test_USAA_parser():
    log = """posted,,12/31/2018,,DGX - PHILADELPHIA       PHILADELPHIA PA,General Merchandise,-9.35
             posted,,5/1/2018,,USAA CREDIT CARD PAYMENT SAN ANTONIO  TX,Credit Card Payments,--507.83
             posted,,2/6/2018,,AVE C MKT xxxxxx0062     TROY         MI,Restaurants/Dining,-1.79"""
    tempfile = 'test_USAA_parser.csv'
    assert True == True

    # TODO check out pytest fixtures, see if it can be used to make temporary files


if __name__ == "__main__":
    test_USAA_parser()
