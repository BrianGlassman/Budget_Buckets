# -*- coding: utf-8 -*-
"""
Created on Sat Aug  1 16:32:23 2020

@author: Brian Glassman


#%%
Parses raw input files from banks, Venmo, etc.

Parsers are classes to make it easier to group information and helper methods
"""

#%% Definitions
import csv
from abc import ABC

from transaction import RawTransaction, Transaction

def parse(infile, parser, *args, **kwargs):
    """Run the provided parsing logic on the given file."""
    return parser.parse(infile, *args, **kwargs)


class USAA_parser(ABC):
    """Namespace for handling csv output files from USAA.

    Files have no header, csv with lines of the form:
    [string status],,[string MM/DD/YYYY],,[string description],[string category],[float value]

    Note: value is -[float] for expenses, --[float] for income
    """

    fieldnames = ['status', 'idk', 'date', 'custom_desc', 'auto_desc', 'auto_cat', 'value']

    @classmethod
    def parse(cls, infile, account):
        """Parse the given file, tagging transactions as the given account.

        Args:
            infile (file path): path to the input file to open
            account (string): the account to tag transactions as
        """
        assert account is not None, "USAA parser called without specifying account"
        with open(infile, 'r') as f:
            return [cls.make_transaction(line, account) for line in csv.DictReader(f, cls.fieldnames)]

    @classmethod
    def make_transaction(cls, line, account):
        """Convert a line from the input file into a RawTransaction object.

        Args:
            line (dict of strings): the line read in from the input file
            account (string): the account to tag transactions as
        Returns:
            raw_transaction object
        """
        # Make empty fields None
        line = {k:(v if v else None) for k,v in line.items()}
        desc = cls.get_desc(**line)
        value = cls.format_value(**line)
        return RawTransaction(date = line['date'],
                              desc = desc,
                              auto_cat = line['auto_cat'],
                              value = value,
                              account = account)

    @staticmethod
    def get_desc(custom_desc, auto_desc, **excess):
        """Combines custom_desc and auto_desc
        Args:
            custom_desc (str): manually input description
            auto_desc (str): automatically generated description
            **excess: unused fields from the input line
        Returns:
            one desc filled: return unchanged
            both filled: "[auto] --- [custom]"
            neither: empty string
        """
        if custom_desc and auto_desc:
            return f"{auto_desc} --- {custom_desc}"
        elif custom_desc and (not auto_desc): return custom_desc
        elif (not custom_desc) and auto_desc: return auto_desc
        elif (not custom_desc) and (not auto_desc): return ""
        else: raise RuntimeError # Shouldn't be possible

    @staticmethod
    def format_value(value, **excess):
        """Properly format the value
        Args:
            value (str): monetary value. Positive values may have "--" prefix
            **excess: unused fields from the input line
        Returns:
            value as a float
        """
        if value.startswith('--'):
            return float(value[2:])
        else:
            return float(value)


def venmo_parser():
    """
    Two line header of the form:
     Username,ID,Datetime,Type,Status,Note,From,To,Amount (total),Amount (fee),Funding Source,Destination,Beginning Balance,Ending Balance,Statement Period Venmo Fees,Year to Date Venmo Fees,Disclaimer
    [string],,,,,,,,,,,,$[float],,,,

    Data lines of the form:
    ,[int],[YYYY-MM-DDTHH:MM:SS],[string type],[string status],[string description],[string payer],[string payee],[+-] $[float],,[string funding source],,,,,,

    Four line footer of the form:
    ,,,,,,,,,,,,,$[float],,,
    ,,,,,,,,,,,,,,$[float],,
    ,,,,,,,,,,,,,,,$[float],
    ,,,,,,,,,,,,,,,,"[multi-line text disclaimer]"

    """


def cash_parser():
    pass


class excel_parser(ABC):
    """Namespace for parsing my existing Excel files"""

    xlfile = 'ref_Budget Buckets.xlsm'

    fieldnames = {'Date': 'date',
                  'Custom Description': 'custom_desc',
                  'Auto Description': 'auto_desc',
                  'USAA Category': 'auto_cat',
                  'Raw Amount': 'value',
                  'Account': 'account',}

    @classmethod
    def parse(cls, infile=xlfile, accounts=None):
        """Parse the file.

        Args:
            infile (file path): path to the input file to open
        """
        import openpyxl
        wb = openpyxl.load_workbook(filename=cls.xlfile, read_only=False, data_only=True)

        out = []
        try:
            # Different years are stored in different sheets
            sheets = ['Log 2018']
            for ws in (wb[sheetname] for sheetname in sheets):
                # Each row after the header is one transaction
                # Skip header in row 1

                # Format as {field : 0-indexed column number}, only keep used columns
                header = {cell.value:c for c,cell in enumerate(ws[1]) if cell.value in cls.fieldnames.keys()}
                # no min_col because that would mess up the indexing
                max_col = max(header.values()) + 1  # max_col parameter uses 1-index

                for raw_row in ws.iter_rows(min_row=2, max_col=max_col, values_only=True):
                    row = {cls.fieldnames[field]:raw_row[c] for field,c in header.items()}

                    # Filter by account
                    if accounts is not None and row['account'] in accounts:
                        out.append(cls.make_transaction(row))
        finally:
            # Openpyxl doesn't release open read-only workbooks
            wb.close()

        return out


    @classmethod
    def make_transaction(cls, line):
        """Convert a line from the input file into a RawTransaction object.

        Args:
            line (dict of strings): the line read in from the input file
        Returns:
            raw_transaction object
        """
        desc = cls.get_desc(**line)
        value = cls.format_value(**line)
        return RawTransaction(date = line['date'],
                              desc = desc,
                              auto_cat = line['auto_cat'],
                              value = value,
                              account = line['account'])

    @staticmethod
    def get_desc(custom_desc, auto_desc, **excess):
        """Combines custom_desc and auto_desc
        Args:
            custom_desc (str): manually input description
            auto_desc (str): automatically generated description
            **excess: unused fields from the input line
        Returns:
            one desc filled: return unchanged
            both filled: "[auto] --- [custom]"
            neither: empty string
        """
        if custom_desc and auto_desc:
            return f"{auto_desc} --- {custom_desc}"
        elif custom_desc and (not auto_desc): return custom_desc
        elif (not custom_desc) and auto_desc: return auto_desc
        elif (not custom_desc) and (not auto_desc): return ""
        else: raise RuntimeError  # Shouldn't be possible

    @staticmethod
    def format_value(value, **excess):
        """Properly format the value
        Args:
            value (str): monetary value. Positive values may have "--" prefix
            **excess: unused fields from the input line
        Returns:
            value as a float
        """
        if isinstance(value, str) and value.startswith('--'):
            return float(value[2:])
        else:
            return float(value)

class excel_cat_parser(excel_parser):
    fieldnames = {'Date': 'date',
                  'Custom Description': 'custom_desc',
                  'Auto Description': 'auto_desc',
                  'Raw Amount': 'value',
                  'Account': 'account',
                  'My Category': 'category',
                  'Split_raw': 'split'}

    @staticmethod
    def get_recurrence(split, **excess):
        # Apply recurrence
        if split == 1:
            # Single transaction
            return None
        else:
            # Recurring transaction
            # Have to do some fuzzy logic, this won't be in production
            # NOTE: initial count is reduced by one to account for first instance
            if split == 14:
                # Salary
                return '14d' # Apply over 14 days, don't normalize
            elif 28 <= split <= 31:
                # Monthly
                return '1m' # Apply over 1 month, don't normalize
            elif 365 <= split <= 366:
                # Yearly
                return '1y' # Apply over 1 year, don't normalize

    @classmethod
    def make_transaction(cls, line):
        """Convert a line from the input file into a Transaction object.

        Args:
            line (dict of strings): the line read in from the input file
        Returns:
            raw_transaction object
        """
        desc = cls.get_desc(**line)
        value = cls.format_value(**line)
        rec = cls.get_recurrence(**line)
        return Transaction(date = line['date'],
                           desc = desc,
                           category = line['category'],
                           value = value,
                           recurrence = rec)

#%% Testing
def test_USAA_parser():
    log = """posted,,12/31/2018,,DGX - PHILADELPHIA       PHILADELPHIA PA,General Merchandise,-9.35
             posted,,5/1/2018,,USAA CREDIT CARD PAYMENT SAN ANTONIO  TX,Credit Card Payments,--507.83
             posted,,2/6/2018,,AVE C MKT xxxxxx0062     TROY         MI,Restaurants/Dining,-1.79"""
    tempfile = 'test_USAA_parser.csv'
    assert True == True

    # TODO check out pytest fixtures, see if it can be used to make temporary files


if __name__ == "__main__":
    test_USAA_parser()
