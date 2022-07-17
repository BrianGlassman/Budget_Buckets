# -*- coding: utf-8 -*-
"""
Created on Sun Jun 28 14:29:34 2020

@author: Brian Glassman
"""

"""Script for applying categories from my original Excel file to the new logs
Purely for testing and rapid development, shouldn't be used in final version
Note: creates the "categorized" log"""

import datetime

# from temp_excel2log import logfile
from bank2log import outfile as logfile

#%% Read in the Excel file
import openpyxl
# NOTE: HAVE TO BE CAREFUL ABOUT INDEXING. sheet.cell() uses 1-index, but
# tuples created from rows/columns use 0-index

xlfile = '../ref_Budget Buckets.xlsm'

# Fields used in production version
real_fields = ('Raw ID', 'Date', 'Custom Description', 'Auto Description',
               'Auto Category', 'Raw Amount', 'Account')
# Fields only used in this temp version
temp_fields = ('Split_raw', 'My Category')
# Combine real and temp fields
fields = real_fields + temp_fields

wb = openpyxl.load_workbook(filename=xlfile, read_only=False, data_only=True)
try:
    # Different years are stored in different sheets
    sheets = ['Log 2018']
    for ws in (wb[sheetname] for sheetname in sheets):
        # Each row after the header is one transaction
        # Skip header in row 1
        # Column A (1) is the Raw ID
        # Columns B-H (2-8) are original data, columns H+ are post-processed
        # Columns K (11) and M (13) are used to categorize and split
        # Column N (14) is the Account
        # for raw_row in ws.iter_rows(min_row=2, max_col=12, values_only=True): # 2019 or 2020

        # Format as {field : 0-indexed column number}, only keep used columns
        header = {cell.value:c for c,cell in enumerate(ws[1]) if cell.value in fields}
        min_col = min(header.values()) + 1 # min/max column uses 0-index
        max_col = max(header.values()) + 1 # min/max column uses 1-index

        excel = []
        for raw_row in ws.iter_rows(min_row=2, min_col=min_col, max_col=max_col, values_only=True):
            row = {field:raw_row[c] for field,c in header.items()}

            if row['Account'] != 'Secured Credit':
                continue

            row['Raw ID'] = None if row['Raw ID'] is None else int(row['Raw ID'])
            row['Date'] = row['Date'].date()
            excel.append(row)
finally:
    # Openpyxl doesn't release open read-only workbooks
    wb.close()

# Sort chronologically (earliest at the top, latest at the bottom)
excel = sorted(excel, key=lambda x: x['Date'])

#%% Split based on day, do other stuff

def get_desc(raw):
    c = raw['Custom Description']
    a = raw['Auto Description']
    if c and a:
        return f"{a} --- {c}"
    elif c and (not a):
        return c
    elif (not c) and a:
        return a
    elif (not c) and (not a):
        return ""
    else:
        raise ValueError

log = []
for mod_id, raw_trans in enumerate(excel):
    # TODO for now, mod_id = raw_id
    # Production version will include splitting transactions and shifting start times, which will change mod_id

    # Rename fields for later convenience (also avoid pass-by-reference)
    trans = {'date': raw_trans['Date'], # transaction date
             'value': raw_trans['Raw Amount'], # transaction value, negative for expenses, positive for income
             'desc': get_desc(raw_trans), # transaction description
             'raw_id': raw_trans['Raw ID'], # ID of parent object from raw log, to preserve backwards traceability
             'mod_id': mod_id, # unique ID within the modded log, to store link information in files
                            # Also used to link recurring transactions
             'category': raw_trans['My Category'], # transaction category
             'recurrence': '?', # How to apply recurrence (gets filled in below)
                 # "None" means this is a one-time transaction, to be applied immediately
                 # Lower case is applied without shifting
                 # Upper case normalizes
                 # d/D = days
                 # m/M = months
                 # y/Y = years
             }

    # Apply recurrence
    rec = raw_trans['Split_raw']
    if rec == 1:
        # Single transaction
        trans['recurrence'] = None
    else:
        # Recurring transaction
        # Have to do some fuzzy logic, this won't be in production
        if rec == 14:
            # Salary
            rec = '14d' # Apply over 14 days, don't normalize
        elif 28 <= rec <= 31:
            # Monthly
            rec = '1m' # Apply over 1 month, don't normalize
        elif 365 <= rec <= 366:
            # Yearly
            rec = '1y' # Apply over 1 year, don't normalize

        trans['recurrence'] = rec

    # Add to the log
    log.append(trans)

# Sort chronologically (earliest at the top, latest at the bottom)
log = sorted(log, key=lambda x: x['date'])

#%% Create the JSON log
import json

jstr = json.dumps(log, default=str, indent=2)
print(jstr)

json.dump(log, open('mod_log.txt', 'w'), indent=2, default=str)

#%% NOTES ON LOG FORMAT
"""
Save as a dict ordered by date
    Easier to handle recurring events this way
"""
